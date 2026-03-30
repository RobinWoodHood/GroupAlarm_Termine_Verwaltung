from __future__ import annotations

import logging
from pathlib import Path
from typing import Callable, List

from openpyxl import Workbook
from openpyxl.styles import Alignment

from .appointment import Appointment
from .importer_token import ImporterToken

logger = logging.getLogger(__name__)

COLUMNS = [
    "name",
    "description",
    "startDate",
    "endDate",
    "organizationID",
    "labelIDs",
    "isPublic",
    "reminder",
    "notificationDate",
    "feedbackDeadline",
    "timezone",
    "groupalarm_id",
    "ga_importer_token",
    "feedback_positive",
    "feedback_negative",
    "feedback_no_response",
]


def _format_datetime(dt, tz_name: str) -> str:
    """Internal helper for `format_datetime`."""
    if dt is None:
        return ""
    from dateutil import tz as dateutil_tz
    target_tz = dateutil_tz.gettz(tz_name)
    if target_tz:
        dt = dt.astimezone(target_tz)
    return dt.isoformat()


def export_appointments(
    appointments: List[Appointment],
    output_path: Path,
    timezone: str = "Europe/Berlin",
    user_name_resolver: Callable[[int], str] | None = None,
) -> Path:
    """Execute `export_appointments`."""
    if not appointments:
        raise ValueError("No appointments to export")

    wb = Workbook()
    ws = wb.active
    ws.title = "Appointments"
    ws.append(COLUMNS)

    for appt in appointments:
        clean_desc, token = ImporterToken.strip_from_text(appt.description)
        token = token or ""
        feedback_groups: dict[int, list[str]] = {0: [], 1: [], 2: []}
        for participant in appt.participants or []:
            status = participant.get("feedback", 0) or 0
            status_key = status if status in (0, 1, 2) else 0
            user_id = participant.get("userID")
            if user_name_resolver and user_id is not None:
                name = user_name_resolver(user_id)
            elif participant.get("name"):
                name = str(participant.get("name"))
            else:
                name = f"User #{user_id}" if user_id is not None else "Unbekannt"

            comment = participant.get("feedbackMessage")
            if comment:
                name = f"{name} ({comment})"
            feedback_groups[status_key].append(name)

        row = [
            appt.name,
            clean_desc,
            _format_datetime(appt.startDate, timezone),
            _format_datetime(appt.endDate, timezone),
            appt.organizationID,
            ",".join(str(lid) for lid in appt.labelIDs) if appt.labelIDs else "",
            appt.isPublic,
            appt.reminder if appt.reminder is not None else "",
            _format_datetime(appt.notificationDate, timezone),
            _format_datetime(appt.feedbackDeadline, timezone),
            appt.timezone,
            appt.id if appt.id is not None else "",
            token,
            "\n".join(feedback_groups[1]),
            "\n".join(feedback_groups[2]),
            "\n".join(feedback_groups[0]),
        ]
        ws.append(row)
        feedback_start_column = len(COLUMNS) - 2
        for column in range(feedback_start_column, len(COLUMNS) + 1):
            ws.cell(row=ws.max_row, column=column).alignment = Alignment(wrap_text=True)

    wb.save(output_path)
    logger.info("Exported %d appointments to %s", len(appointments), output_path)
    return output_path
