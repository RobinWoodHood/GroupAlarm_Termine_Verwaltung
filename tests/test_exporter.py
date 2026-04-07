"""Tests for exporter module — export appointments to .xlsx."""
import pytest
from datetime import datetime, timezone

from framework.appointment import Appointment
from framework.exporter import export_appointments, COLUMNS


def _make_appt(
    id_=1,
    name="Test",
    description="desc [GA-IMPORTER:aabbccdd|20260101120000|abcd]",
    participants=None,
):
    """Internal helper for `make_appt`."""
    return Appointment(
        name=name,
        description=description,
        startDate=datetime(2026, 3, 22, 10, 0, tzinfo=timezone.utc),
        endDate=datetime(2026, 3, 22, 14, 0, tzinfo=timezone.utc),
        organizationID=100,
        id=id_,
        labelIDs=[1, 2],
        isPublic=True,
        timezone="Europe/Berlin",
        participants=participants or [],
    )


def test_export_creates_xlsx_with_16_columns(tmp_path):
    """Test `export_creates_xlsx_with_16_columns` behavior."""
    path = tmp_path / "test.xlsx"
    result = export_appointments([_make_appt()], path)
    assert result == path
    assert path.exists()

    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    assert headers == COLUMNS
    assert len(headers) == 16


def test_export_contains_groupalarm_id_and_token(tmp_path):
    """Test `export_contains_groupalarm_id_and_token` behavior."""
    path = tmp_path / "test.xlsx"
    export_appointments([_make_appt(id_=42)], path)

    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb.active
    row = [cell.value for cell in ws[2]]
    # groupalarm_id is column 12 (0-indexed: 11)
    assert row[11] == 42
    # ga_importer_token is column 13 (0-indexed: 12)
    assert "[GA-IMPORTER:" in str(row[12])


def test_export_empty_list_raises(tmp_path):
    """Test `export_empty_list_raises` behavior."""
    path = tmp_path / "test.xlsx"
    with pytest.raises(ValueError, match="No appointments"):
        export_appointments([], path)


def test_export_multiple_appointments(tmp_path):
    """Test `export_multiple_appointments` behavior."""
    path = tmp_path / "test.xlsx"
    appts = [_make_appt(id_=i, name=f"Appt {i}") for i in range(1, 4)]
    export_appointments(appts, path)

    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb.active
    # header + 3 data rows
    assert ws.max_row == 4


def test_export_appointment_without_token(tmp_path):
    """Test `export_appointment_without_token` behavior."""
    path = tmp_path / "test.xlsx"
    appt = _make_appt(description="no token here")
    export_appointments([appt], path)

    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb.active
    row = [cell.value for cell in ws[2]]
    # ga_importer_token should be empty/None (openpyxl stores "" as None)
    assert not row[12]


def test_export_feedback_columns_use_names_comments_and_linebreaks(tmp_path):
    """Test `export_feedback_columns_use_names_comments_and_linebreaks` behavior."""
    path = tmp_path / "test.xlsx"
    appt = _make_appt(
        participants=[
            {"userID": 10, "feedback": 1, "feedbackMessage": "komme spaeter"},
            {"userID": 11, "feedback": 1},
            {"userID": 12, "feedback": 2, "feedbackMessage": "verhindert"},
            {"userID": 13, "feedback": 0},
        ]
    )
    name_map = {10: "Alice Example", 11: "Bob Example", 12: "Cara Example", 13: "Dan Example"}
    export_appointments([appt], path, user_name_resolver=lambda uid: name_map[uid])

    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb.active
    row = [cell.value for cell in ws[2]]

    assert row[13] == "Alice Example (komme spaeter)\nBob Example"
    assert row[14] == "Cara Example (verhindert)"
    assert row[15] == "Dan Example"


def test_export_label_names_with_resolver(tmp_path):
    """Labels are exported as names when label_name_resolver is provided."""
    path = tmp_path / "test.xlsx"
    label_map = {1: "Zugführer", 2: "Bereitschaft"}
    export_appointments(
        [_make_appt()],
        path,
        label_name_resolver=lambda lid: label_map.get(lid, str(lid)),
    )

    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb.active
    row = [cell.value for cell in ws[2]]
    # labelIDs column (index 5)
    assert row[5] == "Zugführer,Bereitschaft"


def test_export_label_ids_without_resolver(tmp_path):
    """Labels are exported as numeric IDs when no label_name_resolver is provided."""
    path = tmp_path / "test.xlsx"
    export_appointments([_make_appt()], path)

    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb.active
    row = [cell.value for cell in ws[2]]
    # labelIDs column (index 5)
    assert row[5] == "1,2"
